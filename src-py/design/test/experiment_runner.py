import logging
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

from matplotlib import pyplot as plt
import pandas as pd

from design.core.models.network_data import NetworkData
from design.core.optimizer import ShippingNetworkOptimizer
from src.utils.config import Config
from src.utils.visualization import NetworkVisualizer


class ExperimentRunner:
    def _record_results_to_excel(self, config, results):
        """将优化结果记录到Excel文件"""
        try:
            excel_path = "results/results.xlsx"
            sheet_name = f"Instance{config.Instance}"
            
            # 安全获取结果数据，不存在则返回None
            def get_result(method, metric):
                try:
                    return results.get(method, {}).get("design_solution", {}).get(f"total_{metric.lower()}")
                except:
                    return None

            # 获取当前时间
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            # 打开Excel文件或创建新文件
            try:
                wb = load_workbook(excel_path)
            except FileNotFoundError:
                wb = Workbook()
                wb.remove(wb.active)  # 删除默认创建的工作表
            
            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                # 添加表头
                headers = [
                    ["Instance", "P", "K", "R", "T", "S",
                     "Gurobi", "Gurobi", "Gurobi",
                     "Cost", "Cost", "Cost",
                     "Utility", "Utility", "Utility",
                     "Demand", "Demand", "Demand",
                     "备注"],
                    ["Instance", "P", "K", "R", "T", "S",
                     "Cost", "Utility", "Demand",
                     "Cost", "Utility", "Demand",
                     "Cost", "Utility", "Demand",
                     "Cost", "Utility", "Demand",
                     ""]
                ]
                for header in headers:
                    sheet.append(header)
            else:
                sheet = wb[sheet_name]
            
            # 准备新行数据
            new_row = [
                config.Instance,
                config.P,
                config.K,
                config.R,
                config.T,
                config.seed,
                get_result("Gurobi", "cost"),  # Gurobi Cost
                get_result("Gurobi", "utility"),  # Gurobi Utility
                get_result("Gurobi", "capatured_demand"),  # Gurobi Demand
                get_result("Cost", "cost"),  # Cost Cost
                get_result("Cost", "utility"),  # Cost Utility
                get_result("Cost", "capatured_demand"),  # Cost Demand
                get_result("Utility", "cost"),  # Utility Cost
                get_result("Utility", "utility"),  # Utility Utility
                get_result("Utility", "capatured_demand"),  # Utility Demand
                get_result("Demand", "cost"),  # Demand Cost
                get_result("Demand", "utility"),  # Demand Utility
                get_result("Demand", "capatured_demand"),  # Demand Demand
                f"{current_time}"  # 备注
            ]
            sheet.append(new_row)
            
            # 保存文件
            wb.save(excel_path)
            logging.info(f"结果已成功记录到{excel_path}的{sheet_name}工作表")
            
        except Exception as e:
            logging.error(f"记录结果到Excel时出错: {str(e)}", exc_info=True)

    def run_optimization(self,
                         config: Config = None,
                         ports_df: pd.DataFrame = None,
                         P: int = 30,
                         K:int = 0,
                         R: int = 5,
                         seed:int = 0):
        """运行单次优化任务"""
        self.config = config
        self.config.setup_logger(ods= K, seed= seed)
        
        logging.info("==================================================================")
        logging.info(f"开始运行优化任务: P={P}, K={K}, S={seed}")

        try:
            # 创建数据
            logging.info("正在创建数据...")
            # 创建网络数据
            network_data = NetworkData(
                config= config,
                ports_data= ports_df,
                num_ports=config.P,
                num_ods=config.K,
                num_routes=config.R,
                random_seed=seed
            )
            # 绘制节点分布
            networkVisualizer = NetworkVisualizer(network_data=network_data)
            networkVisualizer.plot_port_distribution(ports_df= ports_df)
            network_data.print_network_data()

            # 初始化并运行优化器
            logging.info("开始求解模型...")
            optimizer = ShippingNetworkOptimizer(network_data= network_data)
            optimizer.add_test("ALNS", "Cost")
            optimizer.add_test("ALNS", "Utility")
            optimizer.add_test("ALNS", "Demand")
            start_time = time.time()
            results= optimizer.optimize()
            elapsed_time = time.time() - start_time
            logging.info(f"求解完成，用时{elapsed_time:.2f} s")

            # 记录结果到Excel
            self._record_results_to_excel(config, results)

            logging.info("==================================================================")

            return results

        except Exception as e:
            logging.error(f"运行优化任务时出错: {str(e)}", exc_info=True)
            raise
